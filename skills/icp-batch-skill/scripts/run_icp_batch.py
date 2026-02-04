#!/usr/bin/env python3
"""
Batch ICP query workflow with cache-first strategy.

Steps:
- Extract domains from workbook (default: domains.xlsx, column named '链接').
- Load cache file (icp_results.csv) to reuse successes (status_code=200, code=1).
- Call API for missing/failed entries using AppCode auth.
- Rewrite cache, write success summary, and append two columns (备案主体/备案号) to workbook.
"""

import argparse
import json
import os
import re
import sys
import time
from pathlib import Path
from typing import Dict, List, Tuple, cast

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
import requests


DOMAIN_PATTERN = re.compile(r"[A-Za-z0-9.-]+\.[A-Za-z]{2,}")


def extract_domains(workbook_path: Path, link_header: str = "链接") -> List[str]:
    wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    ws = cast(Worksheet, wb.active)
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return []

    # Locate column
    col_idx = None
    for idx, val in enumerate(header):
        if val == link_header:
            col_idx = idx
            break
    if col_idx is None:
        col_idx = 0  # fallback to first column

    seen = set()
    domains: List[str] = []
    for row in rows:
        if col_idx >= len(row):
            continue
        cell_val = row[col_idx]
        if cell_val is None:
            continue
        text = str(cell_val).strip()
        m = DOMAIN_PATTERN.search(text)
        domain = m.group(0).lower() if m else text.lower()
        if domain not in seen:
            seen.add(domain)
            domains.append(domain)
    return domains


def load_cache(cache_path: Path) -> Dict[str, Dict[str, str]]:
    if not cache_path.exists():
        return {}
    import csv

    with cache_path.open(encoding="utf-8") as f:
        reader = csv.DictReader(f)
        return {row["domain"]: row for row in reader if "domain" in row}


def parse_success(body: str):
    try:
        data = json.loads(body)
    except Exception:
        return None
    if not isinstance(data, dict) or data.get("code") != 1:
        return None
    return data.get("data") or {}


def call_api(domain: str, appcode: str, host: str, path: str, session: requests.Session):
    url = f"{host}{path}"
    resp = session.get(url, params={"domain": domain}, headers={"Authorization": f"APPCODE {appcode}"}, timeout=10)
    return {
        "domain": domain,
        "status_code": str(resp.status_code),
        "error_header": resp.headers.get("X-Ca-Error-Message", ""),
        "body": resp.text,
    }


def rewrite_cache(cache_path: Path, all_domains: List[str], rows_by_domain: Dict[str, Dict[str, str]]):
    import csv

    fieldnames = ["domain", "status_code", "error_header", "body"]
    with cache_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for dom in all_domains:
            row = rows_by_domain.get(dom)
            if row:
                writer.writerow(row)


def write_success(success_path: Path, rows_by_domain: Dict[str, Dict[str, str]]):
    import csv

    fieldnames = ["domain", "icp_name", "icp_num", "sitename", "service", "status"]
    parsed = []
    for row in rows_by_domain.values():
        if row.get("status_code") != "200":
            continue
        data = parse_success(row.get("body", ""))
            
        if not data:
            continue
        parsed.append({
            "domain": data.get("domain", ""),
            "icp_name": data.get("icp_name", ""),
            "icp_num": data.get("icp_num", ""),
            "sitename": data.get("sitename", ""),
            "service": data.get("service", ""),
            "status": data.get("status", ""),
        })

    with success_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(parsed)


def update_workbook(workbook_path: Path, success_rows: Dict[str, Tuple[str, str]], link_header: str = "链接"):
    wb = openpyxl.load_workbook(workbook_path)
    ws = cast(Worksheet, wb.active)

    max_col = ws.max_column
    subject_col = max_col + 1
    num_col = max_col + 2
    ws.cell(row=1, column=subject_col, value="备案主体")
    ws.cell(row=1, column=num_col, value="备案号")

    # locate column index again
    header_row = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col_idx = None
    for idx, val in enumerate(header_row):
        if val == link_header:
            col_idx = idx
            break
    if col_idx is None:
        col_idx = 0

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if not row:
            continue
        cell_val = row[col_idx].value
        if cell_val is None:
            continue
        text = str(cell_val).strip()
        m = DOMAIN_PATTERN.search(text)
        domain = m.group(0).lower() if m else text.lower()
        info = success_rows.get(domain)
        if not info:
            continue
        row_idx = row[0].row
        if row_idx is None:
            continue
        ws.cell(row=row_idx, column=subject_col, value=info[0])
        ws.cell(row=row_idx, column=num_col, value=info[1])

    wb.save(workbook_path)


def read_appcode_file(path: Path) -> str:
    if not path.exists():
        return ""
    try:
        value = path.read_text(encoding="utf-8").strip()
    except Exception:
        return ""
    return value


def resolve_appcode(cli_value: str) -> str:
    if cli_value:
        return cli_value

    # try current directory
    current = Path.cwd() / "appcode.txt"
    value = read_appcode_file(current)
    if value:
        return value

    # try executable/script directory
    if getattr(sys, "frozen", False):
        base_dir = Path(sys.executable).resolve().parent
    else:
        base_dir = Path(__file__).resolve().parent
    value = read_appcode_file(base_dir / "appcode.txt")
    if value:
        return value

    if sys.stdin and sys.stdin.isatty():
        try:
            return input("请输入AppCode: ").strip()
        except Exception:
            return ""
    return ""


def choose_workbook_via_dialog() -> Path:
    try:
        import tkinter as tk
        from tkinter import filedialog
    except Exception:
        return Path()

    try:
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        path = filedialog.askopenfilename(
            title="选择要处理的 Excel 文件",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        root.destroy()
    except Exception:
        return Path()

    return Path(path) if path else Path()


def resolve_workbook(path_value: str) -> Path:
    workbook = Path(path_value)
    if workbook.exists():
        return workbook

    # if default filename, try next to executable/script
    if path_value == "domains.xlsx":
        if getattr(sys, "frozen", False):
            base_dir = Path(sys.executable).resolve().parent
        else:
            base_dir = Path(__file__).resolve().parent
        alt = base_dir / path_value
        if alt.exists():
            return alt

    dialog_path = choose_workbook_via_dialog()
    if dialog_path.exists():
        return dialog_path

    if sys.stdin and sys.stdin.isatty():
        try:
            prompt = f"未找到 {workbook}，请输入文件路径: "
            value = input(prompt).strip()
        except Exception:
            value = ""
        if value:
            alt = Path(value)
            if alt.exists():
                return alt

    sys.exit(f"Workbook not found: {workbook}")


def should_use_gui(force_gui: bool) -> bool:
    if force_gui:
        return True
    if getattr(sys, "frozen", False):
        if sys.stdout is None or sys.stderr is None:
            return True
    if not sys.stdin or not sys.stdin.isatty():
        return True
    return False


def prompt_appcode_gui() -> str:
    try:
        import tkinter as tk
        from tkinter import simpledialog
    except Exception:
        return ""
    try:
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        value = simpledialog.askstring("输入", "请输入 AppCode:")
        root.destroy()
    except Exception:
        return ""
    return value.strip() if value else ""


def format_seconds(seconds: float) -> str:
    if seconds < 0:
        seconds = 0
    minutes, sec = divmod(int(seconds), 60)
    hours, minutes = divmod(minutes, 60)
    if hours > 0:
        return f"{hours:02d}:{minutes:02d}:{sec:02d}"
    return f"{minutes:02d}:{sec:02d}"


class ProgressUI:
    def __init__(self, total: int, use_gui: bool):
        self.total = total
        self.use_gui = use_gui
        self.root = None
        self.label = None
        self.progress = None
        if use_gui:
            try:
                import tkinter as tk
                from tkinter import ttk
            except Exception:
                self.use_gui = False
                return
            self.root = tk.Tk()
            self.root.title("ICP 批量查询")
            self.root.geometry("420x140")
            self.label = tk.Label(self.root, text="准备中...", anchor="w")
            self.label.pack(fill="x", padx=12, pady=(12, 6))
            self.progress = ttk.Progressbar(self.root, maximum=max(1, total))
            self.progress.pack(fill="x", padx=12, pady=6)
            self.root.update()

    def update(self, current: int, message: str):
        if self.use_gui and self.root and self.label and self.progress:
            self.label.config(text=message)
            self.progress["value"] = min(current, self.total)
            self.root.update()
            return
        print(message)

    def info(self, message: str):
        if self.use_gui:
            try:
                from tkinter import messagebox
                messagebox.showinfo("提示", message)
                return
            except Exception:
                pass
        print(message)

    def error(self, message: str):
        if self.use_gui:
            try:
                from tkinter import messagebox
                messagebox.showerror("错误", message)
                return
            except Exception:
                pass
        print(f"ERROR: {message}")

    def close(self):
        if self.use_gui and self.root:
            try:
                self.root.destroy()
            except Exception:
                pass


def main():
    parser = argparse.ArgumentParser(description="Batch ICP query with cache-first strategy")
    parser.add_argument("--workbook", default="domains.xlsx", help="Workbook to process")
    parser.add_argument("--cache", default="icp_results.csv", help="Cache file path")
    parser.add_argument("--success", default="icp_success.csv", help="Parsed success output")
    parser.add_argument("--host", default="https://domainicp.market.alicloudapi.com", help="API host")
    parser.add_argument("--path", default="/do", help="API path")
    parser.add_argument("--appcode", default=os.getenv("APP_CODE"), help="AppCode (or set APP_CODE env/appcode.txt)")
    parser.add_argument("--sleep", type=float, default=0.1, help="Sleep between API calls (seconds)")
    parser.add_argument("--gui", action="store_true", help="Force GUI mode")
    args = parser.parse_args()

    use_gui = should_use_gui(args.gui)
    ui = ProgressUI(total=1, use_gui=use_gui)

    try:
        appcode = resolve_appcode(args.appcode)
        if not appcode and use_gui:
            appcode = prompt_appcode_gui()
        if not appcode:
            raise RuntimeError("未提供 AppCode")

        workbook_path = resolve_workbook(args.workbook)

        # If cache/success are default names, place them next to workbook
        cache_path = Path(args.cache)
        success_path = Path(args.success)
        if args.cache == "icp_results.csv" and not cache_path.is_absolute():
            cache_path = workbook_path.parent / cache_path.name
        if args.success == "icp_success.csv" and not success_path.is_absolute():
            success_path = workbook_path.parent / success_path.name

        domains = extract_domains(workbook_path)
        ui.update(0, f"已提取域名 {len(domains)} 条")

        cache = load_cache(cache_path)

        to_call = []
        for dom in domains:
            row = cache.get(dom)
            if not row:
                to_call.append(dom)
                continue
            if row.get("status_code") != "200":
                to_call.append(dom)
                continue
            if not parse_success(row.get("body", "")):
                to_call.append(dom)

        ui.update(0, f"需要调用 API {len(to_call)} 次")

        sess = requests.Session()
        errors = 0
        start_time = time.monotonic()
        ui.total = max(1, len(to_call))
        if ui.use_gui and ui.progress:
            ui.progress["maximum"] = ui.total

        for idx, dom in enumerate(to_call, 1):
            try:
                cache[dom] = call_api(dom, appcode, args.host, args.path, sess)
                if cache[dom].get("status_code") != "200":
                    errors += 1
            except Exception as e:
                cache[dom] = {
                    "domain": dom,
                    "status_code": "-1",
                    "error_header": type(e).__name__,
                    "body": str(e),
                }
                errors += 1

            elapsed = time.monotonic() - start_time
            avg = elapsed / idx if idx else 0
            remaining = avg * (len(to_call) - idx)
            msg = f"处理中 {idx}/{len(to_call)} | 已用 {format_seconds(elapsed)} | 预计剩余 {format_seconds(remaining)}"
            ui.update(idx, msg)
            time.sleep(args.sleep)

        # rewrite cache in domain order
        rewrite_cache(cache_path, domains, cache)

        # build success map
        success_map: Dict[str, Tuple[str, str]] = {}
        for dom, row in cache.items():
            if row.get("status_code") != "200":
                continue
            data = parse_success(row.get("body", ""))
            if not data:
                continue
            success_map[dom] = (data.get("icp_name", ""), data.get("icp_num", ""))

        # write success csv
        write_success(success_path, cache)

        # update workbook columns
        update_workbook(workbook_path, success_map)

        summary = (
            f"完成。总计 {len(domains)} 条，API 调用 {len(to_call)} 次，失败 {errors} 次。\n"
            f"结果文件：{success_path}\n缓存文件：{cache_path}"
        )
        ui.info(summary)
    except Exception as exc:
        ui.error(str(exc))
        sys.exit(1)
    finally:
        ui.close()


if __name__ == "__main__":
    main()
